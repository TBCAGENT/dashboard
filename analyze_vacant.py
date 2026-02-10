import openpyxl
import sys
from collections import defaultdict

try:
    excel_file = '/Users/lukefontaine/.openclaw/media/inbound/572763cc-59c6-44d2-98df-605a05210633.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Sheet1']
    
    print('=== EXTRACTING VACANT-UNRENTED PROPERTIES ONLY ===')
    print('(Excluding squatters, current tenants, etc.)')
    print()
    
    vacant_properties = []
    
    for row in range(2, ws.max_row + 1):
        address = ws.cell(row=row, column=1).value or ''
        city = ws.cell(row=row, column=2).value or ''
        zip_code = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=5).value or ''
        rent = ws.cell(row=row, column=6).value
        
        # Clean up zip code
        if zip_code:
            if isinstance(zip_code, float):
                zip_code = str(int(zip_code))
            else:
                zip_code = str(zip_code).split('.')[0].strip()
        else:
            continue
        
        # ONLY include "Vacant-Unrented" properties
        if 'vacant-unrented' in status.lower():
            vacant_properties.append({
                'address': address,
                'city': city,
                'zip': zip_code,
                'status': status,
                'full_address': f'{address}, {city}, MI {zip_code}'
            })
    
    print(f'Found {len(vacant_properties)} VACANT-UNRENTED properties:')
    print()
    print('Address                            City           Zip    Full Address')
    print('-' * 90)
    
    for i, prop in enumerate(vacant_properties, 1):
        addr_short = prop['address'][:34] if len(prop['address']) > 34 else prop['address']
        city_short = prop['city'][:14] if len(prop['city']) > 14 else prop['city']
        print(f'{i:2}. {addr_short:<32} {city_short:<14} {prop["zip"]:<6} {prop["full_address"]}')
    
    # Group by zip code for analysis
    by_zip = defaultdict(list)
    for prop in vacant_properties:
        by_zip[prop['zip']].append(prop)
    
    print(f'\n=== BREAKDOWN BY ZIP CODE ===')
    for zip_code in sorted(by_zip.keys()):
        print(f'\nZIP {zip_code}: {len(by_zip[zip_code])} vacant properties')
        for prop in by_zip[zip_code]:
            print(f'  • {prop["address"]}')
    
    print(f'\n=== ADDRESSES FOR RESEARCH ===')
    print('Copy these addresses for value research:')
    for prop in vacant_properties:
        print(prop['full_address'])

except Exception as e:
    print(f'Error: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)